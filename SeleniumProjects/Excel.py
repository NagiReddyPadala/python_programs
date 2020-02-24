import openpyxl

path = "F:\Office\Practice\PycharmProjects\SeleniumProjects\data.xlsx"

workbook = openpyxl.load_workbook(path)
#sheet = workbook.get_sheet_by_name("Sheet1")
sheet =workbook.active
rows = sheet.max_row
cols = sheet.max_column

#Read
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end = "   ")
    print()

#Write
for r in range(8, 9):
    for c in range(1, 5):
        sheet.cell(r, c).value = "Tesing"

workbook.save(path)